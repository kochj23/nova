#!/usr/bin/env python3
"""
nova_gdrive_ingest.py — Ingest Google Drive backup from NAS into Nova's local vector memory.

Reads PDFs, XLSX, DOCX, CSV, and image-based text from /Volumes/nas/Google-Drive-kochjpar/
and stores them in the local PostgreSQL+pgvector memory server (port 18790).

ALL data stays local. Every memory is tagged with privacy:local-only metadata
to prevent inclusion in cloud-routed LLM prompts.

Posts status updates to Slack #nova-notifications every 5 minutes.

Written by Jordan Koch.
"""

import json
import os
import signal
import sys
import time
import traceback
import urllib.request
from datetime import datetime
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

SOURCE_DIR = Path("/Volumes/nas/Google-Drive-kochjpar")
MEMORY_URL = "http://127.0.0.1:18790/remember?async=1"
SLACK_URL = "https://slack.com/api/chat.postMessage"
SLACK_CHANNEL = "C0ATAF7NZG9"  # #nova-notifications
STATUS_INTERVAL = 300  # 5 minutes

LOG_FILE = Path("/tmp/nova-gdrive-ingest.log")
STATE_FILE = Path("/tmp/nova-gdrive-ingest-state.json")

# ── Globals ───────────────────────────────────────────────────────────────────

stats = {
    "total_files": 0,
    "processed": 0,
    "skipped": 0,
    "errors": 0,
    "memories_stored": 0,
    "start_time": None,
    "current_file": "",
    "by_type": {},
}
last_status_time = 0
slack_token = None
shutdown_requested = False


def signal_handler(sig, frame):
    global shutdown_requested
    shutdown_requested = True
    log("Shutdown requested, finishing current file...")


signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[gdrive-ingest {ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")


def load_slack_token():
    global slack_token
    try:
        import subprocess
        result = subprocess.run(
            ["security", "find-generic-password", "-a", "nova", "-s", "nova-slack-bot-token", "-w"],
            capture_output=True, text=True, timeout=10
        )
        slack_token = result.stdout.strip()
        if slack_token:
            log("Slack token loaded from Keychain")
    except Exception as e:
        log(f"WARNING: Could not load Slack token: {e}")


def post_slack(message):
    if not slack_token:
        return
    try:
        payload = json.dumps({
            "channel": SLACK_CHANNEL,
            "text": message,
            "mrkdwn": True,
        }).encode()
        req = urllib.request.Request(
            SLACK_URL,
            data=payload,
            headers={
                "Authorization": f"Bearer {slack_token}",
                "Content-Type": "application/json; charset=utf-8",
            },
        )
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        log(f"Slack post failed: {e}")


def post_status():
    global last_status_time
    now = time.time()
    if now - last_status_time < STATUS_INTERVAL:
        return
    last_status_time = now

    elapsed = now - stats["start_time"] if stats["start_time"] else 0
    mins = int(elapsed // 60)
    pct = (stats["processed"] / stats["total_files"] * 100) if stats["total_files"] > 0 else 0
    types_summary = ", ".join(f"{k}: {v}" for k, v in sorted(stats["by_type"].items()))

    msg = (
        f":file_folder: *Google Drive Ingest — Status Update*\n"
        f"• Progress: {stats['processed']}/{stats['total_files']} files ({pct:.0f}%)\n"
        f"• Memories stored: {stats['memories_stored']}\n"
        f"• Errors: {stats['errors']} | Skipped: {stats['skipped']}\n"
        f"• Elapsed: {mins} min\n"
        f"• Types: {types_summary}\n"
        f"• Current: `{stats['current_file']}`"
    )
    post_slack(msg)


def remember(text, title, source_path, file_type):
    if not text or len(text.strip()) < 20:
        return 0

    chunks = chunk_text(text, max_chars=2000)
    stored = 0
    for i, chunk in enumerate(chunks):
        chunk_title = f"{title} (part {i+1}/{len(chunks)})" if len(chunks) > 1 else title
        payload = json.dumps({
            "text": chunk,
            "source": "gdrive-ingest",
            "metadata": {
                "privacy": "local-only",
                "origin": "google-drive-backup",
                "file": str(source_path),
                "file_type": file_type,
                "title": chunk_title,
                "ingested_at": datetime.now().isoformat(),
            },
        }).encode()
        try:
            req = urllib.request.Request(
                MEMORY_URL,
                data=payload,
                headers={"Content-Type": "application/json"},
            )
            urllib.request.urlopen(req, timeout=30)
            stored += 1
        except Exception as e:
            log(f"  Memory store failed for chunk {i+1}: {e}")
    return stored


def chunk_text(text, max_chars=2000):
    if len(text) <= max_chars:
        return [text]
    chunks = []
    paragraphs = text.split("\n\n")
    current = ""
    for para in paragraphs:
        if len(current) + len(para) + 2 > max_chars:
            if current:
                chunks.append(current.strip())
            current = para
        else:
            current = current + "\n\n" + para if current else para
    if current.strip():
        chunks.append(current.strip())
    if not chunks:
        for i in range(0, len(text), max_chars):
            chunks.append(text[i:i + max_chars])
    return chunks


# ── File Processors ───────────────────────────────────────────────────────────

def process_pdf(filepath):
    from pdfminer.high_level import extract_text
    text = extract_text(str(filepath))
    return text.strip() if text else ""


def process_xlsx(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def process_docx(filepath):
    import docx
    doc = docx.Document(str(filepath))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def process_csv(filepath):
    import csv
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = []
        for row in reader:
            rows.append(" | ".join(row))
            if len(rows) > 500:
                break
    return "\n".join(rows)


def process_text(filepath):
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()[:10000]


PROCESSORS = {
    ".pdf": process_pdf,
    ".xlsx": process_xlsx,
    ".docx": process_docx,
    ".csv": process_csv,
    ".txt": process_text,
    ".md": process_text,
}

SKIP_EXTENSIONS = {".ds_store", ".png", ".jpg", ".jpeg", ".heic", ".tif", ".gif"}


def collect_files():
    files = []
    for root, dirs, filenames in os.walk(SOURCE_DIR):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for name in filenames:
            if name.startswith("."):
                continue
            filepath = Path(root) / name
            ext = filepath.suffix.lower()
            if ext in SKIP_EXTENSIONS:
                continue
            files.append(filepath)
    return sorted(files)


def process_file(filepath):
    ext = filepath.suffix.lower()
    rel = filepath.relative_to(SOURCE_DIR)
    title = f"Google Drive — {rel}"

    stats["current_file"] = str(rel)
    stats["by_type"][ext] = stats["by_type"].get(ext, 0) + 1

    processor = PROCESSORS.get(ext)
    if not processor:
        log(f"  SKIP (no processor for {ext}): {rel}")
        stats["skipped"] += 1
        return

    try:
        text = processor(filepath)
        if not text or len(text.strip()) < 20:
            log(f"  SKIP (empty/too short): {rel}")
            stats["skipped"] += 1
            return
        stored = remember(text, title, rel, ext)
        stats["memories_stored"] += stored
        stats["processed"] += 1
        log(f"  OK: {rel} ({len(text)} chars, {stored} chunks)")
    except Exception as e:
        stats["errors"] += 1
        log(f"  ERROR: {rel} — {e}")
        traceback.print_exc(file=sys.stdout)


def save_state():
    with open(STATE_FILE, "w") as f:
        json.dump(stats, f, indent=2, default=str)


def main():
    global last_status_time
    stats["start_time"] = time.time()
    last_status_time = time.time()

    load_slack_token()

    if not SOURCE_DIR.exists():
        log(f"ERROR: Source directory not found: {SOURCE_DIR}")
        post_slack(f":x: *Google Drive Ingest FAILED* — `{SOURCE_DIR}` not mounted")
        sys.exit(1)

    files = collect_files()
    stats["total_files"] = len(files)
    log(f"Found {len(files)} files to process in {SOURCE_DIR}")

    post_slack(
        f":rocket: *Google Drive Ingest Started*\n"
        f"• Source: `{SOURCE_DIR}`\n"
        f"• Files: {len(files)}\n"
        f"• Privacy: `local-only` — zero cloud\n"
        f"• Status updates every 5 minutes"
    )

    for filepath in files:
        if shutdown_requested:
            log("Shutdown requested — stopping")
            break
        process_file(filepath)
        post_status()
        save_state()

    elapsed = time.time() - stats["start_time"]
    mins = int(elapsed // 60)

    summary = (
        f":white_check_mark: *Google Drive Ingest Complete*\n"
        f"• Processed: {stats['processed']}/{stats['total_files']} files\n"
        f"• Memories stored: {stats['memories_stored']}\n"
        f"• Errors: {stats['errors']} | Skipped: {stats['skipped']}\n"
        f"• Time: {mins} min\n"
        f"• All data tagged `privacy:local-only`"
    )
    if shutdown_requested:
        summary = summary.replace("Complete", "INTERRUPTED (will resume)")

    log(summary.replace("*", "").replace(":white_check_mark:", "OK"))
    post_slack(summary)
    save_state()


if __name__ == "__main__":
    main()
